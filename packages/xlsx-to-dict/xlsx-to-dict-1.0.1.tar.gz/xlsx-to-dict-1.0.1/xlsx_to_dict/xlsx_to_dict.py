from typing import Optional, Dict, List

import openpyxl
from slugify import slugify


def xlsx_to_dict(filename: str, sheetname: Optional[str]=None) -> List[Dict]:
    fields = []
    data = []
    
    wb = openpyxl.load_workbook(filename, read_only=True)
    ws = wb[sheetname or wb.sheetnames[0]]

    for x, itens in enumerate(ws.iter_rows(values_only=True)):
        if x == 0:
            fields.extend([slugify(item).replace('-', '_') for item in itens])
        else:
            data.append({k: v for k, v in zip(fields, itens)})
    
    wb.close()

    return data
