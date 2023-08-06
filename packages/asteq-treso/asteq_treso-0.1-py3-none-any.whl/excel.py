import os
from openpyxl import load_workbook
import pandas as pd
import json

# ------------------------------------------------- MANAGEMENT FUNCTIONS ----------------------------------------------------

def createExcelIfNotExisting(filename_excel)->bool:
    if not(os.path.isfile(filename_excel)):
        print(f"Excel file '{filename_excel}' doesn't exists, creating a new file...")
        with pd.ExcelWriter(filename_excel,engine='xlsxwriter'):
            pass
        return False
    return True

def createSheetIfNotExisting(filename_excel, sheet_name)->bool:
    wb_readonly = load_workbook(filename_excel, read_only=True)
    if not(sheet_name in wb_readonly.sheetnames) :
        wb_readonly.close()
        print(f"Sheet '{sheet_name}' in '{filename_excel}' doesn't exists, creating a new sheet")
        wb = load_workbook(filename_excel)
        wb.create_sheet(sheet_name)
        wb.save(filename_excel)
        wb.close()
        return False
    return True

#------------------------------------------------------ JSON/EXCEL FUNCTIONS ----------------------------------------------------

def appendJsonToExcel(data,filename_excel,sheet_name):
    # Json to Dataframe
    df = pd.json_normalize(data)

    # Create excel file if not existing
    excelExist = createExcelIfNotExisting(filename_excel)

    # Create sheet if not existing
    sheetExist = createSheetIfNotExisting(filename_excel,sheet_name)

    # Append to Excel File if existing:
    if sheetExist&excelExist:
        with pd.ExcelWriter(filename_excel,mode='a',if_sheet_exists='overlay',engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name,startrow=writer.sheets[sheet_name].max_row,header=None,index=None)
    else :
        with pd.ExcelWriter(filename_excel,engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name,index=None)

def excelToJson(filename_excel,sheet_name,save_in_file:str=None):
    # Get data from .xlsx file
    try :
        excel_data_df = pd.read_excel(filename_excel, sheet_name=sheet_name,header=0)

        json_str = excel_data_df.to_json(orient="records")
        json_dict = json.loads(json_str)

        if (save_in_file != None):
            with open(save_in_file,mode='w') as f:
                json.dump(json_dict,f,indent=4)

        return json.dumps(json_dict, indent=4)
    except :
        return json.dumps([])