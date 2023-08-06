


def cmdLineInput(tips='请输入应用启动类型: '):
    import re
    str = input(tips).lower()
    arrays = re.split(r"[ ]+", str)
    if arrays[len(arrays) - 1] == '':
        del arrays[len(arrays) - 1]
    if len(arrays) == 0:
        return None,(),()
    appType = arrays[0]
    params = []
    for index in range(1, len(arrays)):
        params.append(arrays[index])
    params = []
    for index in range(1, len(arrays)):
        params.append(arrays[index])
    return appType,tuple(params),tuple(arrays)

def getNowTimeStr(timeFormat='%Y-%m-%d %H:%M:%S'):
    '''获取当前时间字符串'''
    import datetime
    return datetime.datetime.strftime(datetime.datetime.now(), timeFormat)

def printFormat(format_str, *args):
    if len(args) > 0:
        print(format_str % args)
    else:
        print(format_str)

def isPortInUse(port, ip='127.0.0.1'):
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        s.connect((ip, port))
        s.shutdown(2)
        return True
    except:
        return False

def validConfigFile(config, func):
    '''效验配置文件'''
    sections = config.sections()
    for section in sections:
        options = config.options(section)
        for option in options:
            value = config.get(section, option)
            new_value = func(value)
            if new_value != value:
                config.set(section, option, new_value)

def listDuplicate(data_list):
    '''列表去重'''
    from functools import reduce
    run_function = lambda x, y: x if y in x else x + [y]
    return reduce(run_function, [[], ] + data_list)

def convertFloatNum(_v):
    if type(_v) == str and len(_v) == 0:
        return 0
    else:
        return float(_v)

def createXLSX(filePath,titleArr,sheetName='Sheet1'):
    try:
        import xlsxwriter
        workbook = xlsxwriter.Workbook(filePath)  # 创建一个Excel文件
        worksheet = workbook.add_worksheet(sheetName)  # 创建一个sheet
        worksheet.write_row(0, 0, titleArr)  # title 写入Excel
        workbook.close()
    except Exception as e:
        print("EXCEL创建失败", e)

def writeXLSX(filePath,rowsData,nextRow=-1,sheetName='Sheet1'):
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(filePath)
        worksheet = workbook.get_sheet_by_name(sheetName)
        if nextRow == -1:
            nextRow = worksheet.max_row + 1  # 获得行数
        rowIndex = 1
        for it in rowsData:
            worksheet.cell(nextRow, rowIndex).value = it
            rowIndex = rowIndex + 1
        workbook.save(filePath)
    except Exception as e:
        print("EXCEL写入失败",e)

def writeXLSXBatch(filePath,rowsDataList,sheetName='Sheet1'):
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(filePath)
        worksheet = workbook.get_sheet_by_name(sheetName)
        startRow = worksheet.max_row + 1  # 获得行数
        for rowsData in rowsDataList:
            rowIndex = 1
            for it in rowsData:
                worksheet.cell(startRow, rowIndex).value = it
                rowIndex = rowIndex + 1
            print("当前写入行(%d),总行数(%d)" % (startRow, len(rowsDataList)))
            startRow=startRow+1
        workbook.save(filePath)
    except Exception as e:
        print("EXCEL批量写入失败",e)

def excelConvertTypeValue(valueType, value):
    # print('excelConvertTypeValue', valueType, type(value), value)
    try:
        # 单元格的ctype属性为0时，对应的python格式为空字符串：''
        if valueType == 0:
            return ''
        # 单元格的ctype属性为1时，对应的python格式为字符串
        elif valueType == 1:
            return value.strip()
        # 单元格的ctype属性为2时，对应的python格式为float和int
        elif valueType == 2:
            if value % 1 == 0.0:
                return str(int(value))
            else:
                return str(value)
        # 单元格的ctype属性为3时，对应的python格式为datetime
        elif valueType == 3:
            import datetime,xlrd
            dateStr = "%.4d-%.2d-%.2d %.2u:%.2u:%.2u" % xlrd.xldate_as_tuple(value, 0)
            return dateStr
        # 单元格的ctype属性为4时，对应的python格式为Bool
        elif valueType == 4:
            return True if value == 1 else False
    except Exception as e:
        # print('excel转换单元格数值错误', valueType, type(value), value, e)
        return ''
    return value

def excelSheetRead(tableName,table,contentList,filterContentList,dict_convert_head = None,filter_rows = None,filter_rows_args=None):
    if table.nrows == 0: return
    # 表头 第一行
    headsRows = table.row_values(0)
    for i in range(1, table.nrows):
        excel_rows = table.row_values(i)
        excel_rows_type = table.row_types(i)
        # 每行数据
        rows_dict = {"excel_line": '%s : %d' % (tableName,i)}
        # 读取行内容
        for j in range(0, len(headsRows)):
            key = headsRows[j]
            if len(str(key)) == 0: continue
            value = excel_rows[j]
            value_type = excel_rows_type[j]
            # 转换excel字段
            rows_dict[key] = excelConvertTypeValue(value_type, value)
        # 转换表头
        if dict_convert_head:
            for headName in dict_convert_head:
                convertHeadName = dict_convert_head[headName]
                if headName in rows_dict:
                    rows_dict[convertHeadName] = rows_dict.pop(headName)  # 字段名转换
                else:
                    rows_dict[convertHeadName] = ''  # 字段不存在,设置默认值
        # 过滤每行字段
        if filter_rows:
            isFilter = False
            try:
                args = (rows_dict,)
                if filter_rows_args:
                    args = args + filter_rows_args
                filterResult = filter_rows(*args)
                if filterResult is not None:
                    if type(filterResult) == bool:
                        isFilter = filterResult
                        if isFilter: rows_dict['filter_reason'] = ''
                    elif type(filterResult) == str:
                        isFilter = True
                        rows_dict['filter_reason'] = filterResult
                    elif type(filterResult) == tuple:
                        isFilter = filterResult[0]
                        rows_dict['filter_reason'] = filterResult[1]
            except Exception as e:
                import traceback
                traceback.print_exc()
                # print('过滤Excel rows,错误: %s' % e, rows_dict)
                isFilter = True
                rows_dict['filter_reason'] = str(e)
            if isFilter:
                filterContentList.append(rows_dict)
                # print('过滤数据',rows_dict)
                continue
        contentList.append(rows_dict)


def excelToBean(excelPath,dict_convert_head = None,filter_rows = None,filter_rows_args=None):
    """
     Excel列表转对象
    :param excelPath: excel全路径
    :return: list[{head:value,head2:value2...}]
    """
    import xlrd
    contentList = []
    filterContentList = []
    try:
        excel = xlrd.open_workbook(excelPath)
        for index in range(0,excel.nsheets):
            table = excel.sheets()[index]
            tableName = excel.sheet_names()[index]
            try:
                excelSheetRead(tableName,table,contentList,filterContentList,dict_convert_head,filter_rows,filter_rows_args)
            except:
                import traceback
                traceback.print_exc()
    except Exception as e:
        import traceback
        traceback.print_exc()
        print('读取excel文件,失败',excelPath,e)
    return contentList, filterContentList




def objToJson(obj, indent=None, customJSONEncoder=None):
    '''对象转JSON'''
    import json
    from lsptoolbox.common_class import DefaultJSONEncoder
    if not customJSONEncoder: customJSONEncoder = DefaultJSONEncoder
    jsonStr = json.dumps(obj, ensure_ascii=False, cls=customJSONEncoder, indent=indent)
    return jsonStr


def createResult(isSuccess = True, data = None , message = None, toJson = True, error=None):
    '''创建resp固定结构的返回值'''
    dict = {}
    if isSuccess:
        dict['code'] = 200
        dict['message'] = 'SUCCESS'
    else:
        dict['code'] = 0
        dict['message'] = 'FAIT'
    if message is not None:
        dict['message'] = message
    if data is not None:
        dict['erpdata'] = data
    if error is not None:
        dict['code'] = -200
        dict['message'] = 'ERROR'
        dict['error'] = str(error)
    if toJson:
        return objToJson(dict)
    return dict

def outToFileJsonFormat(dataBean,filePath):
    """json格式化保存对象到指定文件"""
    with open(filePath, 'w+',encoding='utf-8') as f:
        f.write(objToJson(dataBean,1))

def urlToLocalFile(url,localDirPath):
    '''下载url变成本地文件'''
    try:
        import requests,os
        if not os.path.exists(localDirPath):
            os.makedirs(localDirPath)
        resp = requests.get(url=url)
        fname = os.path.basename(url)
        fullPath = localDirPath + fname
        with open(fullPath, "wb+") as locFile:
            locFile.write(resp.content)
            return fullPath
    except Exception as e:
        print("URL转换本地文件",e)
    return None

def removeSymbol(sentence,remove_chars=None):
    '''过滤特殊符号'''
    if remove_chars is None:
        remove_chars = r'[·’!"\#$%&\'()（）＃！（）*+,-./:;<=>?\@，：?￥★、…．＞【】［］《》？“”‘’\[\\]^_`{|}~]+ⅡII'
    list = []
    for c in sentence.strip():
        list.append(c.strip(remove_chars))
    return ''.join(list)

def isEnglishWords(strs):
    import string
    for i in strs:
        if i not in string.ascii_lowercase + string.ascii_uppercase:
            return False
    return True

def delEnglishChar(str):
    '''移除英文字母'''
    nstr = ''
    for it in str:
        asciiCode = ord(it)
        if 65 <= asciiCode <= 90:continue
        if 97 <= asciiCode <= 122:continue
        if 0 <= asciiCode <= 9:continue
        nstr = nstr + it
    return nstr

def removeEnglishStr(str):
   # 判断是否全英文药品
   if isEnglishWords(str) is False:
       # 移除 英文字母干扰
       return delEnglishChar(str)
   return str

def recodeProcessPID(recode_pid_file):
    '''记录进程PID'''
    try:
        recode_pid = None
        import os
        if os.path.exists(recode_pid_file):
            # 存在已记录的PID文件
            with open(recode_pid_file, 'w') as f:
                try:
                    recode_pid = int(f.readline())
                except:
                    pass
        if recode_pid is not None:
            # 判断记录pid是否正在运行
            import psutil
            if psutil.pid_exists(recode_pid):
                return False
            # 记录当前程序的PID到文件
        with open(recode_pid_file, 'w') as f:
            f.write(str(os.getpid()))
            return True
    except Exception as e:
        print('尝试记录当前程序PID失败',recode_pid_file,e)
    return False

def writeFileContent(path,content):
    '''文件内容写入'''
    try:
        with open(path, 'w',encoding='utf-8') as f:
            f.write(str(content))
    except Exception as e:
        print('写入文件(%s)失败' % path, e)

def readFileContent(path,convertFunc=None):
    '''文件内容读取'''
    content = None
    try:
        import os
        if os.path.exists(path):
            with open(path, 'r',encoding='utf-8') as f:
                content = f.read()
            if convertFunc and content and len(content)>0:
                content = convertFunc(content)
    except Exception as e:
        print('读取文件(%s)失败' % path , e)
    return content

def calMillisecondsTimeDiff(access_start, access_end):
    '''计算毫秒数'''
    return int((access_end -access_start).seconds * 1000 + (access_end -access_start).microseconds / 1000)

def hommizationTimeDiff(access_start,access_end):
    '''格式化 时分秒'''
    delta = access_end - access_start
    seconds = int(round(delta.total_seconds()))
    minutes, seconds = divmod(seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return "{:d}:{:02d}:{:02d}".format(hours, minutes, seconds)

def hommizationMillis(ts):
    millis = int(ts % 1000)
    seconds = int((ts/1000)%60)
    minutes = int((ts/(1000*60))%60)
    hours = int((ts/(1000*60*60))%24)
    day = int((ts/(1000*60*60*24))%360)
    info = ''
    if day>0: info+= str(day)+'天'
    if hours>0: info+= str(hours)+'小时'
    if minutes>0: info+= str(minutes)+'分钟'
    if seconds>0: info+= str(seconds)+'秒'
    if millis>0: info+= str(millis)+'毫秒'
    return  info

def jsonHandlePrev(obj):
    if isinstance(obj, dict):
        for k in list(obj):
            v = obj[k]
            if isinstance(k, bytes):
                obj[str(k, 'utf-8')] = obj.pop(k)
            jsonHandlePrev(v)
    if isinstance(obj, list):
        for it in obj:
            jsonHandlePrev(it)

def datetimeType(v):
    import datetime
    return datetime.datetime.strptime(v, "%Y-%m-%d %H:%M:%S") if type(v) == str else v

def stringSimilarityScore(str1, str2):
    import difflib, math
    score = difflib.SequenceMatcher(None, str(str1), str(str2)).quick_ratio()
    return math.ceil(score * 100)

def convertPercentage(float_val,percent=2):
    return str(format(float_val, '.'+str(percent)+'%'))

def decimalRound_old(value, digit):
    ''' float 四舍五入 '''
    from decimal import Decimal

    if value is None or str(value) == '':
        raise ValueError(' decimalRound value is '' or None ')

    result = str(Decimal(str(value)))

    if float(value) < 0:
        result = result[1:]
        if result != '':
            indexDec = result.find('.')
            if indexDec > 0:
                decimal = result[indexDec + 1:]
                decimalCount = len(decimal)
                if decimalCount > digit:
                    xiaoshu = result[indexDec + digit + 1]  # 第digit+1位小数
                    if int(xiaoshu) > 4:
                        result = str(float(value) * -1 + pow(10, digit * -1))
                        # 存在进位的可能，小数点会移位
                        indexDec = result.find('.')
                        result = result[:indexDec + digit + 1]
                    else:
                        result = result[:indexDec + digit + 1]
                else:
                    lens = digit - len(result[indexDec:]) + 1
                    for i in range(lens):
                        result += '0'
        result = float(result) * -1
    else:
        indexDec = result.find('.')
        if indexDec > 0:
            decimal = result[indexDec + 1:]
            decimalCount = len(decimal)
            if decimalCount > digit:
                xiaoshu = result[indexDec + digit + 1]  # 第digit+1位小数
                if int(xiaoshu) > 4:
                    result = str(float(value) + pow(10, digit * -1))
                    # 存在进位的可能，小数点会移位
                    indexDec = result.find('.')
                    result = result[:indexDec + digit + 1]
                else:
                    result = result[:indexDec + digit + 1]
            else:
                lens = digit - len(result[indexDec:]) + 1
                for i in range(lens):
                    result += '0'
    return float(result)

def decimalRound(value, digit=4):
    ''' dicimal 四舍五入 '''
    if value is None or str(value) == '':
        raise ValueError(' decimalRound value is '' or None ')
    import decimal
    from decimal import Decimal
    decimal.getcontext().rounding = "ROUND_HALF_UP"
    _digit =  '0.' + '0'*digit  if digit>0 else '0'
    return Decimal(str(value)).quantize(Decimal(_digit))

def roundStr(value,digit=4):
    format = '%.'+str(digit)+'f'
    return format % decimalRound(value,digit)

def div(a, b):
    ''' 除法'''
    try:
        from decimal import Decimal
        return float( Decimal(str(a)) / Decimal(str(b) ) )
    except:
        pass
    return 0

def mut(a, b):
    '''乘法'''
    try:
        from decimal import Decimal
        return float( Decimal(str(a)) * Decimal(str(b) ) )
    except:
        pass
    return 0

def add(a, b):
    '''加法'''
    try:
        from decimal import Decimal
        return float( Decimal(str(a)) + Decimal(str(b) ) )
    except:
        pass
    return 0

def sub(a, b):
    '''减法'''
    try:
        from decimal import Decimal
        return float( Decimal(str(a)) - Decimal(str(b) ) )
    except:
        pass
    return 0

def div_round(a, b, digit=4):
    return decimalRound(div(a, b), digit)

def mul_round(a, b, digit=4):
    return decimalRound(mut(a, b), digit)

def add_round(a, b, digit=4):
    return decimalRound(add(a, b), digit)

def sub_round(a, b, digit=4):
    return decimalRound(sub(a, b), digit)

def div_round_float(a, b, digit=4):
    return float(decimalRound(div(a, b), digit))

def mul_round_float(a, b, digit=4):
    return float(decimalRound(mut(a, b), digit))

def add_round_float(a, b, digit=4):
    return float(decimalRound(add(a, b), digit))

def sub_round_float(a, b, digit=4):
    return float(decimalRound(sub(a, b), digit))

def amt_divide(amt, num):
    '''计算金额及除不尽的余数'''
    if float(amt) <= 0 or int(num) <= 0: return 0.0, 0.0
    uprice = div_round(amt, num)
    indiv = sub(amt, mut(uprice, num))
    return uprice, indiv


def strConvertMD5(s):
    try:
        import hashlib
        myMd5 = hashlib.md5()
        myMd5.update(s.encode('utf-8'))
        myMd5_Digest = myMd5.hexdigest()
        return str(myMd5_Digest).upper()
    except: pass
    return s

def strToDatetime(datetimeStr):
    try:
        import datetime
        return datetime.datetime.strptime(datetimeStr, '%Y-%m-%d %H:%M:%S')
    except:
        return None

def strToDate(datetimeStr):
    try:
        import datetime
        return datetime.datetime.strptime(datetimeStr, '%Y-%m-%d')
    finally:
        return None

def join_string(*args):
    '''sql语句 字符串合并'''
    if len(args) <1: return None
    curStr = args[0]
    index = 1
    while(index<len(args)):
        curStr = args[index] % curStr
        index+=1
    return curStr


def getModuleFileSpecSuffixVars(module,suffix):
    '''获取指定模块文件的指定后缀的变量'''
    varList = []
    var_dict = dir(module)
    for varname in var_dict:
        if varname.endswith(suffix):
            var = getattr(module,varname)
            # print(' 获取varname ', varname, var)
            varList.append(var)
    return varList

def getMultipleModuleFileSpecSuffixVars(moduleList,suffix):
    '''获取多个模块文件的指定后缀的变量'''
    varList = []
    for module in moduleList:
        tempList = getModuleFileSpecSuffixVars(module,suffix)
        varList.extend(tempList)
    return varList


def socket_send_message(socket_conn,msg):
    '''socket 发送内容 处理粘包'''
    import struct
    bytes = msg.encode("utf-8")
    bytes_len = len(bytes)
    socket_conn.send(struct.pack('i', bytes_len))
    socket_conn.send(bytes)

def socket_receive_message(socket_conn):
    '''socket 接收内容 处理粘包'''
    import struct
    data = socket_conn.recv(4)
    len = struct.unpack('i', data)[0]
    bytes = socket_conn.recv(len)
    content = bytes.decode("utf-8")
    return content