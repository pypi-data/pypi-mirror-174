"""
A script that converts a network Excel file into the request format.

The Excel file has to be in one of the supported formats, which examples can
be found in the data_formats directory.
"""
import xlrd
import openpyxl
import pathlib

from contextlib import suppress


def convert(excel_file):
    for subclass in Converter.__subclasses__():
        converter = subclass(excel_file)

        return converter.convert()


class Converter:
    def __init__(self, excel_file):
        self._excel_file = excel_file

        self._file_extension = pathlib.Path(excel_file).suffix
        if self._file_extension not in [".xls", ".xlsx"]:
            raise FormatUnsupportedError(f'File must be of format ".xls" or ".xlsx".')
        
        if self._file_extension=='.xls':
            self._file = xlrd.open_workbook(excel_file)
        else:
            self._file = openpyxl.load_workbook(excel_file)

    def _get_columns(self, sheet_name, start=0):
        if self._file_extension=='.xls':
            sheet = self._file.sheet_by_name(sheet_name) # xlrd style
            for colx in range(start, sheet.ncols):
                yield sheet.col_values(colx)[1:]   # ipysheet didn't allow editable column names,
                                                    # so added an index row to sheets and slicing it off here
        else:
            sheet = self._file[sheet_name] # openpyxl style
            for i,colx in enumerate(sheet.columns): # Different writing with openpyxl
                if i<start: continue; # Skip until start line is reached (generator can't skip earlier)
                
                column = [c.value if c.value is not None else '' for c in colx]
                if all([c=='' for c in column]): # openpyxl captures rows and rows of useless NaN.
                    continue;
                yield column[1:]

    def _get_capacities(self):
        raise NotImplementedError
    def _get_links(self):
        raise NotImplementedError
    def _get_network(self):
        raise NotImplementedError

    def convert(self):
        network_request = {
            'capacities': self._get_capacities(),
            'links': self._get_links(),
            'network': self._get_network()
        }

        return network_request


class NewFormatConverter(Converter):
    def _get_columns(self, sheet_name, start=1):
        return super()._get_columns(sheet_name, start=start)

    def _get_capacities(self):
        capacities = []
        for column in self._get_columns('Capacities'):
            (name, units, item_type, options, lower_bound, upper_bound) = column

            if name not in ['#', '']: # to get ipysheet to work
                capacity = {
                    'name': name,
                    'units': units,
                    'type': item_type
                }

                if lower_bound != '' or upper_bound != '':
                    capacity['bounds'] = {}

                    with suppress(ValueError):
                        capacity['bounds']['lower'] = int(lower_bound)
                    with suppress(ValueError):
                        capacity['bounds']['upper'] = int(upper_bound)

                capacities.append(capacity)

        return capacities

    def _get_links(self):
        links = []
        for column in self._get_columns('Network links'):
            (link_id, start_id, end_id, link_type, length, capacity, voltage, electrical_resistance,
             electrical_reactance, total_thermal_loss, total_pressure_loss, operating_temperature) = column

            if link_id not in ['#', '']: # to get ipysheet to work
                with suppress(ValueError):
                    link_id = int(link_id)
                with suppress(ValueError):
                    start_id = int(start_id)
                with suppress(ValueError):
                    end_id = int(end_id)
                with suppress(ValueError):
                    link_type = str(link_type)
                with suppress(ValueError):
                    length = float(length)
                with suppress(ValueError):
                    voltage = float(voltage)
                with suppress(ValueError):
                    electrical_resistance = float(electrical_resistance)
                with suppress(ValueError):
                    electrical_reactance = float(electrical_reactance)
                with suppress(ValueError):
                    total_thermal_loss = float(total_thermal_loss)
                with suppress(ValueError):
                    total_pressure_loss = float(total_pressure_loss)
                with suppress(ValueError):
                    operating_temperature = float(operating_temperature)

                link ={
                    'id': link_id,
                    'start_id': start_id,
                    'end_id': end_id,
                    'type': link_type,
                    # 'link_proportional_cost': link_proportional_cost, uncomment if want to set link cost in the excel file.
                    # add 'link_cost' in the column names
                    'length': length,
                    'voltage': voltage,
                    'resistance': electrical_resistance,
                    'reactance': electrical_reactance,
                    'total_thermal_loss': total_thermal_loss,
                    'total_pressure_loss': total_pressure_loss,
                    'operating_temp': operating_temperature,
                }

                try:
                    capacity = float(capacity)
                except ValueError:
                    capacity = str(capacity)

                link['capacity'] = capacity

                links.append(link)

        return links

    def _get_network(self):
        if self._file_extension=='.xls':
            sheet = self._file.sheet_by_name('Network') # xlrd style
            return{
                'fixed_network_investment_cost': float(sheet.cell(1, 1).value),
                'link_proportional_cost': float(sheet.cell(2, 1).value), 
            } # shifting the sheets one row below required to change the row indices here
        else:
            sheet = self._file["Network"] # openpyxl style
            return{
                'fixed_network_investment_cost': float(sheet.cell(2, 2).value),
                'link_proportional_cost': float(sheet.cell(3, 2).value), 
            } # shifting the sheets one row below required to change the row indices here

        





