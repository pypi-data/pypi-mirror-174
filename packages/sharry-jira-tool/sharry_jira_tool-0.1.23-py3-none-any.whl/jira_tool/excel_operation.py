# -*- coding: utf-8 -*-
"""
This module offers a set of operations that user can modify their excel files.
"""
import logging
import os
import pathlib
import warnings

try:
    from importlib.resources import files
except ImportError:
    from importlib_resources import files

import openpyxl

from .excel_defination import *
from .milestone import *
from .priority import *
from .sprint_schedule import *
from .story import *

__all__ = [
    "read_excel_file",
    "output_to_excel_file",
    "output_to_csv_file",
    "process_excel_file",
]

# Currently, the openpyxl package will report an obsolete warning.
warnings.simplefilter(action="ignore", category=UserWarning)


def read_excel_file(
    file: str,
    excel_defination: ExcelDefination,
    sprint_schedule: SprintScheduleStore,
) -> tuple:
    """
    Read and parse the excel file

    :parm file:
        The excel file that you want to read

    :parm excel_defination:
        The excel column defination which is imported from the :py:class:`ExcelDefination`

    :parm sprint_schedule:
        The priority mapping for the :py:class:`Milestone` object.

    :return:
        A :py:class:`tuple` object which contains a list of column name and a list of :py:class:`Story`.
    """

    if file is None or not pathlib.Path(file).is_absolute():
        raise ValueError(f"The input excel file is invalid.")

    if not pathlib.Path(file).exists():
        raise ValueError(f"The input excel file: {file} cannot be found.")

    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    columns = []

    max_column_index = chr(65 + sheet.max_column - 1)

    start_column = "A1"
    end_column = f"{max_column_index}1"

    for cells in sheet[start_column:end_column]:
        if len(columns) > 0:
            break
        for cell in cells:
            columns.append(cell.value)

    start_cell = "A2"
    end_cell = f"{max_column_index}{sheet.max_row}"
    rows = sheet[start_cell:end_cell]

    stories = []

    excel_defination_columns = excel_defination.get_columns()
    storyFactory = StoryFactory(excel_defination_columns)

    for row in rows:
        if _should_skip(row):
            continue

        story: Story = storyFactory.create_story()
        for column_index in range(len(row)):
            column = excel_defination_columns[column_index]
            story.set_value(column[2], column[1], row[column_index].value)
            if column[2] is Milestone:
                story[column[1]].calc_priority(sprint_schedule)
        stories.append(story)

    wb.close()
    return (columns, stories)


def _should_skip(row: list) -> bool:
    if len(row) == 0:
        return True
    else:
        first_cell_value = row[0].value
        if first_cell_value is None or len(str(first_cell_value)) == 0:
            return True
    return False


def output_to_csv_file(
    file: str,
    stories: "list[Story]",
    over_write: bool = True,
):
    if file is None or not pathlib.Path(file).is_absolute():
        raise ValueError(f"The file is invalid.")

    if not pathlib.Path(file).exists():
        if over_write is True:
            os.remove(file)
        else:
            raise ValueError(f"The csv file: {file} is already exist.")

    with open(file, mode="w") as csv_file:
        separator = "-" * 300
        for story in stories:
            csv_file.write(f"{separator}\n")
            csv_file.write(str(story))


def output_to_excel_file(
    file: str,
    stories: "list[Story]",
    excel_defination: ExcelDefination,
    columns_in_excel: "list[str]" = None,
    over_write: bool = True,
):
    """
    Generate excel file

    :parm file:
        Output excel file name including the path

    :parm stories:
        A list of :py:class:`Story` which need to be wrote to the excel

    :parm excel_defination:
        The excel column defination which is imported from the :py:class:`ExcelDefination`

    :parm columns_in_excel:
        Using separate column names instead of importing from the :py:class:`ExcelDefination`. Usually, it comes from the input excel file.

    :parm over_write:
        Whether or not the exist output file will be over-write.
    """
    if file is None or not pathlib.Path(file).is_absolute():
        raise ValueError(f"The output file name is invalid.")

    if pathlib.Path(file).exists():
        if over_write is True:
            os.remove(file)
        else:
            raise ValueError(f"The output excel file: {file} is already exist.")

    wb = openpyxl.Workbook()
    sheet = wb.active

    excel_defination_columns = excel_defination.get_columns()

    # Use original excel column name first.
    columns = columns_in_excel
    if columns is None:
        columns = [
            column_name for _, column_name, _, _, _, _, _ in excel_defination_columns
        ]

    for column_index in range(len(columns)):
        cell = sheet.cell(row=1, column=column_index + 1)
        cell.value = columns[column_index]

    if stories is not None:
        for row_index in range(len(stories)):
            for column_index, column_name, _, _, _, _, _ in excel_defination_columns:
                cell = sheet.cell(row=row_index + 2, column=column_index)
                cell.value = stories[row_index].format_value(column_name)

    wb.save(file)
    wb.close()


def process_excel_file(
    input_file: str,
    output_file: str,
    sprint_schedule_config: str = None,
    excel_defination_config: str = None,
    over_write: bool = True,
):
    """
    Sort the excel file and output the result

    :parm input_file:
        The excel file need to be sorted. (Absolute path only)

    :parm output_file:
        The sorted excel file location. (Absolute path only)

    :parm sprint_schedule_config:
        The JSON file which contains the priority list to calculate the :py:class:`Milestone`

    :parm excel_defination_config:
        The JSON file which contains the input excel file's structure.

    :parm over_write:
        Whether or not the exist output file will be over-write.
    """
    sprint_schedule = SprintScheduleStore()
    if sprint_schedule_config is None:
        sprint_schedule.load(
            files("jira_tool.assets").joinpath("sprint_schedule.json").read_text()
        )
        print("Loading default sprint schedule...")
    else:
        sprint_schedule.load_file(sprint_schedule_config)

    excel_defination = ExcelDefination()
    if excel_defination_config is None:
        excel_defination.load(
            files("jira_tool.assets").joinpath("excel_defination.json").read_text()
        )
        print("Loading default excel defination...")
    else:
        excel_defination.load_file(excel_defination_config)

    excel_columns, stories = read_excel_file(
        input_file, excel_defination, sprint_schedule
    )

    if stories is None:
        print("There are no stories inside the excel file.")
        return

    sort_strategy_priorities = excel_defination.get_sort_strategy_priorities()

    for sort_strategy in sort_strategy_priorities:
        if sort_strategy.lower() in "InlineWeights".lower():
            stories = sorted(stories, reverse=True)
        elif sort_strategy.lower() in "Sort".lower():
            sort_stories_by_order(stories, excel_defination)
        elif sort_strategy.lower() in "RaiseRanking".lower():
            stories = sort_stories_by_raise_ranking(stories, excel_defination)

    output_to_excel_file(
        output_file, stories, excel_defination, excel_columns, over_write
    )

    print("%s has been saved.", output_file)
