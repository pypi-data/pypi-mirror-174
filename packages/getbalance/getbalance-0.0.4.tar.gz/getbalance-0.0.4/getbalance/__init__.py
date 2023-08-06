from openpyxl.styles import colors, PatternFill, Font, Color
from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal
import requests
import random
import shutil
import time
import json
import sys
import re
import os


def run():
    BLOCK_EXPLORER = {'ETH' : { 'link':'https://ethblockexplorer.org/api/', 
                            'tx_keys': ['unconfirmedTxs', 'txs'],
                            'tx_id' : 'txids'}, 
                    'ETC' : {'link':'https://etcblockexplorer.com/api/', 
                            'tx_keys': ['unconfirmedTxs', 'txs'],
                            'tx_id' : 'txids'},
                    'BTC' : {'link':'https://bitcoinblockexplorers.com/api/', 
                            'tx_keys': ['unconfirmedTxApperances', 'txApperances'],
                            'tx_id' : 'transactions'}, 
                    'BCH' : {'link':'https://bchblockexplorer.com/api/', 
                            'tx_keys': ['unconfirmedTxApperances', 'txApperances'],
                            'tx_id' : 'transactions'}, 
                    'LTC' : {'link':'https://litecoinblockexplorer.net/api/', 
                            'tx_keys': ['unconfirmedTxApperances', 'txApperances'],
                            'tx_id' : 'transactions'}, 
                    'DASH': {'link':'https://dashblockexplorer.com/api/', 
                            'tx_keys': ['unconfirmedTxApperances', 'txApperances'],
                            'tx_id' : 'transactions'}}

    HEADERS = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36" ,
        'referer':'https://www.google.com/'
    }

    def is_valid_xlm_address(address):
        link = f'https://horizon.stellar.org/accounts/{address.strip()}/transactions'
        is_valid = requests.get(link, headers=HEADERS)
        
        if is_valid.status_code != 200:
            return False
            
        return True    
        
    def xlm_chain(address, balance, cursor=''):
        link = f'https://horizon.stellar.org/accounts/{address.strip()}/transactions?limit=200'
        if cursor:
            link = f'{link}&cursor={cursor}'
        txs = requests.get(link, headers=HEADERS)
        content = json.loads(txs.content)
        
        cursor = re.search(r'cursor=(\d+)', content['_links']['next']['href']).group(1)
        
        txs = content['_embedded']['records']
        for tx in txs:
            tx_time = int(datetime.strptime(tx['created_at'].split('T')[0], "%Y-%m-%d").timestamp())
            if tx_time > input_date:
                continue
            tx_link = f'https://horizon.stellar.org/transactions/{tx["id"]}/operations'
            tx = requests.get(tx_link, headers=HEADERS)
            tx_data = json.loads(tx.content)
            tx_data = tx_data['_embedded']['records']
            for tx_datum in tx_data:
                if 'starting_balance' in tx_datum:
                    balance += Decimal(tx_datum['starting_balance'])
                else:
                    if 'from' in tx_datum and tx_datum['from'] == address:
                        balance -= Decimal(tx_datum['amount'])
                    elif 'to' in tx_datum and tx_datum['to'] == address:
                        balance += Decimal(tx_datum['amount'])
        
        if len(txs) == 0:
            return '', balance

        return cursor, balance    
        
    def get_xlm_balance(address):
        balance = 0
        cursor = ''
        cursor, balance = xlm_chain(address, balance)
        while cursor:
            cursor, balance = xlm_chain(address, balance, cursor)
        
        return balance

    def is_valid_xrp_address(address):
        link = f'https://api.xrpscan.com/api/v1/account/{address.strip()}/transactions'
        is_valid = requests.get(link, headers=HEADERS)
        
        if is_valid.status_code != 200:
            return False
            
        content = json.loads(is_valid.content)
        
        if len(content['transactions']) == 0:
            return False

        return True

    def xrp_chain(address, balance, marker=''):
        link = f'https://api.xrpscan.com/api/v1/account/{address.strip()}/transactions'
        if marker:
            link = f'{link}?marker={marker}'
        txs = requests.get(link, headers=HEADERS)
        
        if txs.status_code != 200:
            return '', balance
            
        content = json.loads(txs.content)
        
        txs = content['transactions']
        for tx in txs:
            tx_time = int(datetime.strptime(tx['date'].split('T')[0], "%Y-%m-%d").timestamp())
            if tx_time > input_date:
                continue
            
            currency = tx['Amount']['currency']
            if currency == 'XRP':
                value = tx['Amount']['value']
                if tx['Destination'] == address:
                    balance += value
                else:
                    balance -= value 

        marker = ''
        if 'marker' in content:
            marker = content['marker']
            
        return marker, balance

    def get_xrp_balance(address):
        balance = 0
        marker = ''
        marker, balance = xrp_chain(address, balance)
        while marker:
            marker, balance = xrp_chain(address, balance, marker)
        
        return Decimal(balance) / Decimal(1e6)

    def get_explorer_address_page_count(address, asset):
        link = f'{asset["link"]}address/{address.strip()}'
        is_valid = requests.get(link, headers=HEADERS)
        
        if is_valid.status_code != 200:
            return 0
            
        content = json.loads(is_valid.content)
        
        utxs, txs = asset['tx_keys']
        
        if content[utxs] == 0 and content[txs] == 0:
            return 0
        
        return content['totalPages']

    def get_eth_vin_value(address, tx_data, vin):
        if address.upper() in [*map(lambda x:x.upper(), vin['addresses'])]:
            return Decimal(tx_data['value']) + Decimal(tx_data['fees'])
            
        return 0
        
    def get_eth_vout_value(address, vout):
        if address.upper() in [*map(lambda x:x.upper(), vout['addresses'])]:
            return Decimal(vout['value'])
            
        return 0
        
    def get_btc_vin_value(address, vin):
        if address in vin['addresses']:
            return Decimal(vin['value'])
            
        return 0
        
    def get_btc_vout_value(address, vout):
        if address in vout['scriptPubKey']['addresses']:
            return Decimal(vout['value'])
        
        return 0

    def get_block_explorer_balance(address, asset, page_count):
        current_balance = 0
        is_eth = False
        address = address.strip()

        for page in range(1, page_count+1):
            addr_link = f'{asset["link"]}address/{address}?page={page}'
            addr = requests.get(addr_link, headers=HEADERS)
            
            if addr.status_code != 200:
                return 'error'
                
            content = json.loads(addr.content)
            
            txs = content[asset['tx_id']]
            
            for tx in txs:
                tx_link = f'{asset["link"]}tx/{tx.strip()}'
                tx = requests.get(tx_link, headers=HEADERS)
                
                if tx.status_code != 200:
                    return 'error'
                
                tx_data = json.loads(tx.content)
                
                is_eth = 'ethereumSpecific' in tx_data

                tx_time = tx_data['blockTime'] if is_eth else tx_data['blocktime'] 
                
                if tx_time > input_date:
                        continue
                
                for vin in tx_data['vin']:
                    value = get_eth_vin_value(address, tx_data, vin) if is_eth else get_btc_vin_value(address, vin)
                    current_balance -= value

                for vout in tx_data['vout']:              
                    value = get_eth_vout_value(address, vout) if is_eth else get_btc_vout_value(address, vout)
                    current_balance += value
             
        if is_eth:
            current_balance /= Decimal(1e18)
            
        return current_balance

    INPUT_PATH = 'input/'
    OUTPUT_PATH = 'output/'

    input_date = int(datetime.strptime(input('Input Date (m/d/yyyy): '), "%m/%d/%Y").timestamp()) + 86399

    _, _, files = next(os.walk(INPUT_PATH))

    for file in files:
        shutil.copy(f'{INPUT_PATH}{file}', OUTPUT_PATH)
        work_book = load_workbook(f'{OUTPUT_PATH}{file}', data_only=True)

        for worksheet_idx in range(len(work_book.worksheets)):
            address_idx = None
            currency_idx = None
            address_balance_idx = None
            status_idx = None
            work_book.active = worksheet_idx
            work_sheet = work_book.active
            print(f'Working with {file} - {work_sheet.title}...')
           
            for column_idx, column_name in enumerate(work_sheet[1]):
                if column_name.value == 'Address':
                    address_idx = column_idx
                elif column_name.value == 'Currency':
                    currency_idx = column_idx
                elif column_name.value == 'Address Balance':
                    address_balance_idx = column_idx
                elif column_name.value == 'Status':
                    status_idx = column_idx

            row_count = work_sheet.max_row
            
            def invalidate(row, col):
                work_sheet[row][col].value = 'Null'
                work_sheet[row][col].fill = PatternFill(start_color='F2D3D7', end_color='F2D3D7', fill_type = 'solid')
                work_sheet[row][col].font = Font(color='9C0039')            
                
            def validate(row, col, addr):
                work_sheet[row][col].value = 'Verified'
                work_sheet[row][col].fill = PatternFill(start_color='C3ECCB', end_color='C3ECCB', fill_type = 'solid')
                work_sheet[row][col].font = Font(color='006100')            
           
            for row_idx in range(2, row_count + 1):
                print(f'Working with row #{row_idx}/{row_count}...')
                asset = work_sheet[row_idx][currency_idx].value.upper()
                address = work_sheet[row_idx][address_idx].value                
                
                work_sheet[row_idx][address_balance_idx].number_format = 'General'
                
                if asset in BLOCK_EXPLORER:
                    page_count = get_explorer_address_page_count(address, BLOCK_EXPLORER[asset])
                    if page_count:
                        validate(row_idx, status_idx, address)
                        work_sheet[row_idx][address_balance_idx].value = get_block_explorer_balance(address, BLOCK_EXPLORER[asset], page_count)
                    else:
                        invalidate(row_idx, status_idx)
                elif asset == 'XRP':
                    if is_valid_xrp_address(address):
                        validate(row_idx, status_idx, address)
                        work_sheet[row_idx][address_balance_idx].value = get_xrp_balance(address)
                    else:
                        invalidate(row_idx, status_idx)
                elif asset == 'XLM':
                    if is_valid_xlm_address(address):
                        validate(row_idx, status_idx, address)
                        work_sheet[row_idx][address_balance_idx].value = get_xlm_balance(address)                
                    else:
                        invalidate(row_idx, status_idx)                
        
        work_book.save(f'{OUTPUT_PATH}{file}')            
    print('Balances calculation complete')    
   